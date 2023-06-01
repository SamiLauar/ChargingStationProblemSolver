import os
import time

import pypsa
import openpyxl
import csv
import pandas as pd
import numpy as np
import ProblemClass as pc
from ProblemClass import Problema, Solucao
from typing import List
from random import randint
import random
from openpyxl.chart import (LineChart, Reference, Series)
from IPython.display import display


def gerarSolucaoAleatoria(dados: Problema):
	n_barras = len(dados.Barras)
	max_cs = dados.MaximoEletropostos
	pot_cs = dados.PotEletroposto
	solucao = [0 for _ in range(0, n_barras)]

	while sum(solucao) < dados.PotDemanda:
		pos_novo_cs = randint(0, n_barras - 1)

		# Adiciona um carregador somente se o numero de carregadores na barra não tiver excedido o limite max_cs
		if solucao[pos_novo_cs] < (pot_cs * max_cs):
			solucao[pos_novo_cs] = solucao[pos_novo_cs] + pot_cs

	sol = Solucao(dados, solucao)
	return sol


def gerarSolucaoAleatoriaZonas(dados: Problema):
	pot_cs = dados.PotEletroposto
	max_cs = dados.MaximoEletropostos
	n_barras = len(dados.Barras)
	solucao = np.zeros((n_barras, len(pot_cs)))

	# Atende a restrição de Eletroposto por Zona
	for zone in dados.Zonas:
		iBus = random.choice(zone) - 1
		iPot = randint(0, len(pot_cs) - 1)
		solucao[iBus, iPot] += 1

	# Atende a Res. de Demanda
	while np.sum(pot_cs * np.sum(solucao, axis=0)) < dados.PotDemanda:
		iBus = randint(0, n_barras - 1)
		iPot = randint(0, len(pot_cs) - 1)
		# Adiciona um carregador somente se o numero de carregadores na barra não tiver excedido o limite max_cs
		if solucao[iBus, iPot] < max_cs:
			solucao[iBus, iPot] += 1

	sol = Solucao(dados, solucao)
	return sol


def NormalizeFitness(solucoes: List[Solucao]):
	# i) encontra-se a maior fitness para ser considerada 1
	fit_pu = solucoes[0].fitness

	for sol in solucoes:
		if sol.fitness > fit_pu:
			fit_pu = sol.fitness

	for sol in solucoes:
		sol.fit_pu = fit_pu
		sol.Fitness_pu = sol.fitness / fit_pu


# def FitnessUpdate(solucoes: List[Solucao], i: int = -1):
# 	if i != -1:
# 		solucoes[i].pypsa_update()
# 		solucoes[i].Fitness_pu = solucoes[i].fitness / solucoes[i].fit_pu
# 	else:
# 		for sol in solucoes:
# 			sol.pypsa_update()
# 			sol.Fitness_pu = sol.fitness / sol.fit_pu


def FitnessCalculate(solucao: np.array, dados_problema: Problema):
	solucao_temp = Solucao(dados_problema, solucao)
	# solucao_temp.tornaFactivel()

	# network_temp = pc.criaNetwork(solucao_temp.MatrizSolucao, dados_problema)
	# network_temp.pf()
	#
	# solucao_temp.Network = network_temp
	# solucao_temp.pypsa_read()

	solucao_temp.pypsa_update()

	return solucao_temp.fitness


def tornaFactivel(dados_problema: Problema, matriz_solucao: np.array):
	# Rest 3: Divisão em Zonas
	boolZonaComEletr = False
	for iZone, zone in enumerate(dados_problema.Zonas):
		for iBus in zone:
			if sum(matriz_solucao[iBus - 1]) > 0:
				boolZonaComEletr = True
		if not boolZonaComEletr:
			iBusInZone = random.choice(zone) - 1
			addEletroposto(matriz_solucao, iBusInZone)

	potInstalada = np.sum(dados_problema.PotEletroposto * np.sum(matriz_solucao, axis=0))
	# Rest 1: Pot Instalada > Demanda
	while potInstalada < dados_problema.PotDemanda:
		addEletroposto(matriz_solucao)


def addEletroposto(matriz_solucao: np.array, iBus: int = -1, jEletr: int = -1):
	# Adiciona um eletroposto em uma barra aleatoria ou especificada
	nBus = matriz_solucao.shape[0]
	nEletr = matriz_solucao.shape[1]
	if iBus == -1:
		iBus = randint(0, nBus - 1)
	if jEletr == -1:
		jEletr = randint(0, nEletr - 1)
	matriz_solucao[iBus][jEletr] += 1


def Bounds(universes: List[Solucao], j: int):
	# Encontra o melhor universo
	maior_j = universes[0].MatrizSolucao[j, 0]
	menor_j = maior_j
	for universe in universes:
		for cs_pot in universe.MatrizSolucao:
			if maior_j < cs_pot:
				maior_j = cs_pot
			if menor_j > cs_pot:
				menor_j = cs_pot

	return [maior_j, menor_j]


# The dominates function is a helper function that checks if solution x dominates solution y, meaning that x is at
# least as good as y in all objectives and strictly better than y in at least one objective. This is a common concept
# in multi-objective optimization, where there is no single best solution, but rather a set of solutions that are all
# considered to be Pareto-optimal.
def BestUniverse(universes: List[Solucao], atualBestUnIndex: int):
	best = universes[atualBestUnIndex]
	for i in range(len(universes)):
		if universes[i].fitness < best.fitness:  # Ou menor, depende se é maximize ou minimize
			best = universes[i]
	return best


def BestUniverseIndex(universes: List[Solucao], atualBestUnIndex: int):
	bestU = universes[atualBestUnIndex]
	index = atualBestUnIndex
	for i in range(len(universes)):
		if universes[i].fitness < bestU.fitness:  # Ou menor, depende se é maximize ou minimize
			bestU = universes[i]
			index = i
	return index


def ParetoDominante(population: List[Solucao]):
	vet_pontos = np.zeros((len(population)))
	for i, candidato in enumerate(population):
		for desafiante in population:
			if (candidato.fitnessA < desafiante.fitnessA) and (candidato.fitnessB > desafiante.fitnessB):
				vet_pontos[i] += 1
	dominante = population[np.argmax(vet_pontos)]
	print(f'Dominou {np.argmax(vet_pontos)}')
	return dominante


def FitnessPrint(solucoes: List[Solucao]):
	print(f'Fitness: ')
	print('[', end='')
	for i, sol in enumerate(solucoes):
		if i != len(solucoes) - 1:
			print(f'{sol.fitness:.2f}', end=', ')
		else:
			print(f'{sol.fitness:.2f}', end='] \n')
	print('')


def FitnessString(solucoes: List[Solucao]):
	# strFitness = f'Fitness;  {solucoes[0].fit_pu:.2f};  \n'
	# for sol in solucoes:
	# 	strFitness += f'{sol.fitness:.2f}; '
	# strFitness += '\n'

	vetFitness = [['Best Fitness: ', round(solucoes[0].fitness, 2)], [round(sol.fitness, 2) for sol in solucoes]]

	return vetFitness


def SolucaoString(sol: Solucao):
	strSolucao = '[ ' + ', '.join([str(bus_pot) for bus_pot in sol.MatrizSolucao]) + ']'
	return strSolucao


def SolucaoVetorExcel(sol: Solucao):
	vetSolucao = [['Solucao: PotTotal ', sol.potTotalInstalada], [cs_pot[0] for cs_pot in sol.MatrizSolucao]]

	return vetSolucao


def SolucaoPrint(sol: Solucao):
	print(f'Solucao: PotTotal = {sol.potTotalInstalada}')
	print('[', end='')
	for i, cs_pot in enumerate(sol.MatrizSolucao):
		if i != len(sol.MatrizSolucao) - 1:
			print(cs_pot, end=', ')
		else:
			print(cs_pot, end=']')
	print('')


def ObjMediaPrint(solucoes: List[Solucao], i: int = -1):
	if i == -1:
		nSol = len(solucoes)

		objPerdas = [sol.obj_Perdas() for sol in solucoes]
		mediaPerdas = sum(objPerdas) / nSol
		objVdev = [sol.obj_Vdev() for sol in solucoes]
		mediaVdev = sum(objVdev) / nSol
		objCustoTotal = [sol.obj_CustoTotal() for sol in solucoes]
		mediaCustoTotal = sum(objCustoTotal) / nSol
		objCoberturaDeTrafego = [sol.obj_CoberturaDeTrafego() for sol in solucoes]
		mediaCoberturaDeTrafego = sum(objCoberturaDeTrafego) / nSol

		print(f'Media Perdas                = {mediaPerdas:.1f}')
		print(f'Media Vdev                  = {mediaVdev:.1f}')
		print(f'Media CustoTotal            = {mediaCustoTotal:.0f}')
		print(f'Media CoberturaDeTrafego    = {mediaCoberturaDeTrafego:.3f}')
	else:
		sol = solucoes[i]
		print(f'obj Perdas              = {sol.obj_Perdas():.1f}')
		print(f'obj Vdev                = {sol.obj_Vdev():.1f}')
		print(f'obj CustoTotal          = {sol.obj_CustoTotal():.0f}')
		print(f'obj CoberturaDeTrafego  = {sol.obj_CoberturaDeTrafego():.3f}')


def ObjMediaString(solucoes: List[Solucao], i: int = -1):
	vetObj = []
	if i == -1:
		nSol = len(solucoes)

		objPerdas = [sol.obj_Perdas() for sol in solucoes]
		mediaPerdas = sum(objPerdas) / nSol
		objVdev = [sol.obj_Vdev() for sol in solucoes]
		mediaVdev = sum(objVdev) / nSol
		objCustoTotal = [sol.obj_CustoTotal() for sol in solucoes]
		mediaCustoTotal = sum(objCustoTotal) / nSol
		objCoberturaDeTrafego = [sol.obj_CoberturaDeTrafego() for sol in solucoes]
		mediaCoberturaDeTrafego = sum(objCoberturaDeTrafego) / nSol

		vetObj.append(['Media obj Perdas', round(mediaPerdas, 1)])
		vetObj.append(['Media obj Vdev', round(mediaVdev, 1)])
		vetObj.append(['Media obj CustoTotal', round(mediaCustoTotal, 0)])
		vetObj.append(['Media obj CoberturaDeTrafego', round(mediaCoberturaDeTrafego, 2)])
	else:
		sol = solucoes[i]
		vetObj.append(['obj Perdas', round(sol.obj_Perdas(), 1)])
		vetObj.append(['obj Vdev', round(sol.obj_Vdev(), 1)])
		vetObj.append(['obj CustoTotal', round(sol.obj_CustoTotal(), 0)])
		vetObj.append(['obj CoberturaDeTrafego', round(sol.obj_CoberturaDeTrafego(), 2)])

	return vetObj


def excelAddRow(data: list, file_name: str, sheet_name: str):
	try:
		path = 'results/' + file_name + '.xlsx'

		# Create a new workbook or load an existing one
		if os.path.exists(path):
			wb = openpyxl.load_workbook(path)
		else:
			wb = openpyxl.Workbook()

		# Check if the sheet already exists
		if sheet_name in wb.sheetnames:
			ws = wb[sheet_name]  # Select the existing sheet
		else:
			ws = wb.create_sheet(title=sheet_name)  # Create a new sheet

		# Append the data to the sheet
		ws.append(data)
		# Save the workbook
		wb.save(path)

		print('Data appended to the Excel file successfully.')
	except Exception as e:
		print(f'An error occurred (at Excel append): {str(e)}')


def excelGraphPlot(data: list, file_name: str, sheet_name: str):
	path = 'results/' + file_name + '.xlsx'

	# Create a new workbook or load an existing one
	if os.path.exists(path):
		wb = openpyxl.load_workbook(path)
	else:
		wb = openpyxl.Workbook()

	# Check if the sheet already exists
	if sheet_name in wb.sheetnames:
		ws = wb[sheet_name]  # Select the existing sheet
	else:
		ws = wb.create_sheet(title=sheet_name)  # Create a new sheet

	# Add headers for the data
	ws['A1'] = 'Iteration'
	ws['B1'] = 'Objective Function'

	# Insert sample data for demonstration purposes
	for row in data:
		ws.append(row)

	# Create a line chart using the data
	chart = LineChart()
	chart.title = 'Objective Function x Iteration'
	chart.x_axis.title = 'Iteration'
	chart.y_axis.title = 'Fitness'

	# xdata = Reference(ws, (1, 1), (len(data) + 1, 1))
	# xdata = Reference(ws, min_col=1, min_row=1, max_row=len(data) + 1)
	ydata = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)  # max_row=len(data) + 1)
	chart.add_data(ydata, titles_from_data=True)
	# Add the chart to the worksheet
	ws.add_chart(chart, 'D1')
	# Save the workbook to a file
	wb.save(path)


def networkPrint(network):
	display(network.branches())
	print('\n')
	display(network.loads)
	print('\n')

	display(network.buses_t.v_mag_pu)
	print('\n')
	display(network.buses_t.v_ang)
	print('\n')
	display(network.buses_t.p)
	print('\n')
	display(network.buses_t.q)


def resultsAnalysis(best_solution: Solucao, population: List[Solucao], resultados: list, dadosProblema: Problema, max_iter,
                    max_tempo, start_time, str_algorithm: str = '', str_parametros: str = ''):
	print('\nFim  ------------------------------- \n')
	tempo = time.time() - start_time
	print(f'Time Elapsed: \033[33m{tempo:.1f}\033[ms')

	# Print no Python Console
	SolucaoPrint(best_solution)
	FitnessPrint(population)
	population_fitness = np.array([sol.fitness for sol in population])
	ObjMediaPrint(population, np.argmin(population_fitness))
	best_fitness = best_solution.fitness
	melhora = (resultados[0][1] - best_fitness) / resultados[0][1] * 100
	print(
			f'Resultado: {resultados[0][1]:.3f} para \033[31m{best_fitness:.3f}\033[m \n   Melhora de \033[34m{melhora:.1f}\033[m%')

	tensoes = [[bus.i, bus.V_pu] for i, bus in enumerate(best_solution.Barras)]
	excelGraphPlot(tensoes, str_algorithm + '_V_lastRunResults',
	               str(round(best_fitness, 3)) + '_pot-' + str(dadosProblema.PotDemanda))
	desvio = [1 - vpu[1] for vpu in tensoes]

	# vet das Colunas para Excel
	vetResultados = [
			str_algorithm,
			dadosProblema.PotDemanda,
			best_solution.potTotalInstalada,
			str(round(best_fitness, 3)).replace(".", ","),
			round(resultados[0][1], 3),
			round(melhora, 1),
			round(tempo, 1),
			round(best_solution.obj_Perdas(), 1),
			round(best_solution.obj_Vdev(), 1),
			round(best_solution.obj_CustoTotal(), 1),
			str(round(best_solution.obj_CoberturaDeTrafego(), 3)).replace(".", ","),
			best_solution.obj_Perdas() + best_solution.obj_Vdev() + best_solution.obj_CustoTotal(),
			sum(desvio) / (len(desvio) - 1),
			sum(abs(L) for L in best_solution.Perdas) * 10 ** 3,
			f'alfa1={dadosProblema.k1} / alfa2={dadosProblema.k2} / c2={dadosProblema.c2}',
			f'tarifa={dadosProblema.TarifaRealPorKWh} / penalidade={dadosProblema.MultaDesvTensao} / max_cs={dadosProblema.MaximoEletropostos}',
			f'PopSize={len(population)}, MaxIter={max_iter}, MaxTempo={max_tempo}',
			str_parametros,
			SolucaoString(best_solution)]

	excelAddRow(vetResultados, str_algorithm + '_Results', 'Results')  # 'pot-' + str(problema_teste.PotDemanda))

	excelGraphPlot(resultados, str_algorithm + '_lastRunResults', 'pot-' + str(dadosProblema.PotDemanda))

	for ind in population:
		vet_FObj = [
				ind.potTotalInstalada,
				round(ind.fitness, 3),
				round(ind.fitnessA, 1),
				round(ind.fitnessB, 3),
				round(ind.obj_Perdas(), 1),
				round(ind.obj_Vdev(), 1),
				round(ind.obj_CustoTotal(), 1),
				SolucaoString(ind)]
		excelAddRow(vet_FObj, str_algorithm + '_Pareto_Results', 'Results')

# # Print no Python Console
# solver.FitnessPrint(population)
# solver.ObjMediaPrint(population)
# print('Best:')
# population_fitness = np.array([sol.fitness for sol in population])
# solver.ObjMediaPrint(population, np.argmin(population_fitness))
# print('')

# # open the file in the append mode
# arqOutput = open('data/All_Results.csv', 'a', newline='')
# writerOutput = csv.writer(arqOutput, delimiter=';')
# writerOutput.writerow(vetResultados)
# # close the file
# arqOutput.close()
